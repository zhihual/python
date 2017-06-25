# -*- coding: UTF-8 -*-  
# 运行接受中文字符


# Step1. 数据分类，不同的号分到不同的sheet里面去。 
#        1. 要发现多少个代码
#        2. 然后保存
# 1. 打开，读取一行 判断是否以及是获取过的代码。
# 2. 若不存在，创建新的sheet,添加新代码到已知sheet名字列表里.
# 3. 保存到sheet里面
# 4. 记录以及获取的代码，
# 5. 重复 #1知道数据源的row已经读完

from openpyxl import load_workbook
from openpyxl import Workbook
import os

src_wb_name = '1.xlsx'

src_wb = load_workbook(src_wb_name)

active_src_ws = src_wb.active

src_row_num = len(tuple(active_src_ws.rows))

src_columns_num = len(tuple(active_src_ws.columns))

print("src_row_num=:", src_row_num, "src_columns_num:", src_columns_num)



# 定义一个如何从数据源中获取需要信息的位置
#数据源-交易时间
src_stock_date = 'A'
#数据源-股票号码
src_stock_id_row = 'B'
#数据源-股票中文名称
src_stock_name_row = 'C'
#数据源-操作，买和卖
src_stock_op_row = 'D'
#数据源-单次买卖股票数量
src_stock_deal_num_row = 'E'
#数据源-单次买卖股票价格
src_stock_deal_price_row = 'F'
#数据源-单次操作后，剩余股票数量
src_stock_hold_num = 'H'
#数据源-单次交易券商佣金
src_stock_deal_broker_fee_row = 'J'
#数据源-单次交易印花税
src_stock_deal_tax_row = 'K'
#数据源-单次交易现金变动总额
src_stock_deal_cashinvest = 'I'


dst_wb_name= 'test.xlsx'
dst_wb = Workbook()

active_dst_ws = dst_wb.active

active_dst_ws.title = "week analysis"
active_dst_ws.sheet_properties.tabColor = "1072BA"

# 定义一个空的list,用来存放股票代码
dst_stock_id_list = [] 

# 定义一个
dst_cash_etf_dic = {'511811':'理财金H-赎回', '511991':'华宝添益-赎回', '511881':'银华日利-赎回' }

# 先生成 dst的第一行的标题栏
dst_stock_id_row = 'A'
dst_stock_name_row = 'B'
dst_stock_op_row = 'C'
dst_stock_deal_num_row = 'D'
dst_stock_deal_price_row = 'E'
dst_stock_deal_cashinvest = 'F'
dst_stock_deal_broker_fee_row = 'G'
dst_stock_deal_tax_row = 'H'
dst_stock_hold_num = 'I'
dst_stock_deal_date = 'J'


#定义，一次交易为，为从0买进，到最终卖出到0为一次交易
dst_stock_cycle_mark = 'K'   
#定义，一次交易最终结算的利润统计
dst_stock_intest_mark = 'L'
#定义，日内交易的利润统计
dst_stock_daily_profile_mask = 'M'


#active_dst_ws[dst_stock_id_row+'1'] = '股票代码'
#active_dst_ws[dst_stock_name_row+'1'] = '股票名称'
#active_dst_ws[dst_stock_op_row+'1'] = '买卖'
#active_dst_ws[dst_stock_deal_num_row+'1'] = '成交数量'
#active_dst_ws[dst_stock_deal_price_row+'1'] = '成交价格'
#active_dst_ws[dst_stock_deal_broker_fee_row+'1'] = '券商佣金'
#active_dst_ws[dst_stock_deal_tax_row+'1'] = '印花税'

# 外面应该是个循环
for src_rowidx in range (2, src_row_num+1):
	sheet = active_dst_ws
	# 1. 打开，读取一行 判断是否以及是获取过的代码。
	temp_stock_id = active_src_ws[src_stock_id_row+str(src_rowidx)].value
	temp_stock_id = str(temp_stock_id)
	#print("temp_stock_id:", temp_stock_id)
	
	bIsShuhui = 0
	#此处为每页中，不包含股票中文名称的行，实际上是货币基金的赎回，处理方式为
	#计入本页。
	if temp_stock_id in dst_cash_etf_dic.keys():
		#在字典中
		print(temp_stock_id,src_rowidx,'handle',dst_cash_etf_dic[temp_stock_id] )
		#设置股票中文名字
		active_src_ws[src_stock_name_row+str(src_rowidx)].value = dst_cash_etf_dic[temp_stock_id]
		#设置股票代码， 只为货币基金
		#active_src_ws[src_stock_id_row+str(src_rowidx)].value = int(temp_stock_id) - 1
		#修改本行的股票代码
		temp_stock_id = str(int(temp_stock_id) - 1)
		bIsShuhui = 1
	#2. 若不存在，创建新的sheet,添加新代码到已知sheet名字列表里.
	if temp_stock_id in dst_stock_id_list:
		sheet = dst_wb.get_sheet_by_name(temp_stock_id)
	else:
		#print(temp_stock_id)
		if temp_stock_id:
			sheet = dst_wb.create_sheet(temp_stock_id)
			sheet.sheet_properties.tabColor = "1072BA"
		
			sheet[dst_stock_id_row+'1'] = '股票代码'
			sheet[dst_stock_name_row+'1'] = '股票名称'
			sheet[dst_stock_op_row+'1'] = '买卖'
			sheet[dst_stock_deal_num_row+'1'] = '成交数量'
			sheet[dst_stock_deal_price_row+'1'] = '成交价格'
			sheet[dst_stock_deal_cashinvest+'1'] = '发生金额'
			sheet[dst_stock_deal_broker_fee_row+'1'] = '券商佣金'
			sheet[dst_stock_deal_tax_row+'1'] = '印花税'
			sheet[dst_stock_hold_num+'1'] = '可用余额'
			sheet[dst_stock_deal_date+'1'] = '成交日期'
			sheet[dst_stock_cycle_mark+'1'] = '交易次数'
			sheet[dst_stock_intest_mark+'1'] = '本次利润'
			sheet[dst_stock_daily_profile_mask+'1'] = "日间利润"
			
			
			# 4. 记录以及获取的代码，
			dst_stock_id_list.append(temp_stock_id)
		else:
			#此处为每页中，不包含股票中文名称的行，实际上是货币基金的赎回，处理方式为
			#计入本页。
			#if temp_stock_id in dst_cash_etf_dic.keys():
				#在字典中
			#	print(temp_stock_id,src_rowidx,'handle',dst_cash_etf_dic[temp_stock_id] )
				#设置股票中文名字
			#	active_src_ws[src_stock_name_row+str(src_rowidx)].value = dst_cash_etf_dic[temp_stock_id]
				#设置股票代码， 只为货币基金
			#	active_src_ws[src_stock_id_row+str(src_rowidx)].value = temp_stock_id - 1
			#	src_rowidx = src_rowidx -1
				continue
				
			#else:
			#	print(temp_stock_id,'not in dict', src_rowidx )
			
			
			#此为老的处理方式，直接丢弃。
			#continue
	
	#插入到dst的后面
	sheet_row_num = len(tuple(sheet.rows))
	sheet_row_num = sheet_row_num + 1
	# 3. 保存到sheet里面
	sheet[dst_stock_id_row+str(sheet_row_num)].value = active_src_ws[src_stock_id_row+str(src_rowidx)].value
	sheet[dst_stock_name_row+str(sheet_row_num)].value = active_src_ws[src_stock_name_row+str(src_rowidx)].value
	sheet[dst_stock_op_row+str(sheet_row_num)].value = active_src_ws[src_stock_op_row+str(src_rowidx)].value
	sheet[dst_stock_deal_num_row+str(sheet_row_num)].value = active_src_ws[src_stock_deal_num_row+str(src_rowidx)].value
	sheet[dst_stock_deal_price_row+str(sheet_row_num)].value = active_src_ws[src_stock_deal_price_row+str(src_rowidx)].value
	if bIsShuhui == 1:
		sheet[dst_stock_deal_cashinvest+str(sheet_row_num)].value = -active_src_ws[src_stock_deal_cashinvest+str(src_rowidx)].value
	else:
		sheet[dst_stock_deal_cashinvest+str(sheet_row_num)].value = active_src_ws[src_stock_deal_cashinvest+str(src_rowidx)].value
	sheet[dst_stock_deal_broker_fee_row+str(sheet_row_num)].value = active_src_ws[src_stock_deal_broker_fee_row+str(src_rowidx)].value
	sheet[dst_stock_deal_tax_row+str(sheet_row_num)].value = active_src_ws[src_stock_deal_tax_row+str(src_rowidx)].value
	sheet[dst_stock_hold_num+str(sheet_row_num)].value = active_src_ws[src_stock_hold_num + str(src_rowidx)].value
	sheet[dst_stock_deal_date+str(sheet_row_num)].value = active_src_ws[src_stock_date+str(src_rowidx)].value

	
# 这里，没一个sheet数据都已经整理好了。 开始计算每个sheet的信息，并写入首页

# 1. 根据 dst_stock_id_list 分析每一个sheet，便利，没一个sheet要分析的过程 描述如下：

#    a. 获取有多少row, 从第一row开始。
#    b. 可用余额， 如果余额是0，则跳过此行，直到不为零的行， 标记交易开始
#    c. 根据每行记录金额的变化。 total = previous total + 本行变化
#    d. 若余额从新回到0，则输出本次交易利润，增加交易次数，清零total, 并进入下一行
#    e。标记交易结束， 输出利润到边栏
#    f. 重复直到本页结束

#1 每页都做

exchange_cycle_num = 0
intest = 0.0
daily_prifile = 0.0

#此处开始，进行每页的分析

for anlysis_sheet_name in dst_stock_id_list:
	anlysis_sheet = dst_wb.get_sheet_by_name(anlysis_sheet_name)
	anlysis_sheet_row_num = len(tuple(anlysis_sheet.rows))
	#print(anlysis_sheet_row_num)
	# 开始看每一行的记录
	exchange_cycle_num = 0
	intest = 0.0
	daily_prifile = 0.0
	for anlysis_sheet_row_num in range (2, anlysis_sheet_row_num+1):
		if anlysis_sheet[dst_stock_hold_num+str(anlysis_sheet_row_num)].value == 0:
			#if anlysis_sheet_row_num > 2:
				#intest = intest - anlysis_sheet[dst_stock_deal_num_row+str(anlysis_sheet_row_num)].value * anlysis_sheet[dst_stock_deal_price_row+str(anlysis_sheet_row_num)].value	
				#intest = intest - anlysis_sheet[dst_stock_deal_broker_fee_row+str(anlysis_sheet_row_num)].value
				#intest = intest - anlysis_sheet[dst_stock_deal_tax_row+str(anlysis_sheet_row_num)].value
			#print(intest*-1,'end')
			
			# 记录intest, 清零invest, 增加一次cyclenum
			#exchange_cycle_num = exchange_cycle_num + 1
			#anlysis_sheet[dst_stock_cycle_mark+str(anlysis_sheet_row_num)] = exchange_cycle_num
			#anlysis_sheet[dst_stock_intest_mark+str(anlysis_sheet_row_num)] = intest*-1
			
			#if (intest > -70.0 and (intest*-1) < 70.0):
			#	daily_prifile = daily_prifile+(intest*-1)
			#	anlysis_sheet[dst_stock_daily_profile_mask+str(anlysis_sheet_row_num)] = daily_prifile
			
			#----------------------------------------------------------------------------------
			exchange_cycle_num = exchange_cycle_num + 1
			anlysis_sheet[dst_stock_cycle_mark+str(anlysis_sheet_row_num)] = exchange_cycle_num
			
			if anlysis_sheet_row_num > 2:
				intest += anlysis_sheet[dst_stock_deal_cashinvest+str(anlysis_sheet_row_num)].value
			
			anlysis_sheet[dst_stock_intest_mark+str(anlysis_sheet_row_num)] = intest
			
			if (intest > -70.0 and (intest*-1) < 70.0):
				daily_prifile = intest
				anlysis_sheet[dst_stock_daily_profile_mask+str(anlysis_sheet_row_num)] = daily_prifile
			
			
			intest = 0.0
		else:
			#if anlysis_sheet[dst_stock_op_row+str(anlysis_sheet_row_num)].value == "买入":
			#	intest = intest + anlysis_sheet[dst_stock_deal_num_row+str(anlysis_sheet_row_num)].value * anlysis_sheet[dst_stock_deal_price_row+str(anlysis_sheet_row_num)].value
			#else:
			#	intest = intest - anlysis_sheet[dst_stock_deal_num_row+str(anlysis_sheet_row_num)].value * anlysis_sheet[dst_stock_deal_price_row+str(anlysis_sheet_row_num)].value	
			#intest = intest - anlysis_sheet[dst_stock_deal_broker_fee_row+str(anlysis_sheet_row_num)].value
			#intest = intest - anlysis_sheet[dst_stock_deal_tax_row+str(anlysis_sheet_row_num)].value
			#print(intest, anlysis_sheet_row_num, anlysis_sheet_name)
			
			intest += anlysis_sheet[dst_stock_deal_cashinvest+str(anlysis_sheet_row_num)].value
			
			
		
#验证分析数据是否完全
TotalNum = 0
for anlysis_sheet_name in dst_stock_id_list:
	anlysis_sheet = dst_wb.get_sheet_by_name(anlysis_sheet_name)
	TotalNum = TotalNum + len(tuple(anlysis_sheet.rows)) -1
print(TotalNum+1)	


		
os.remove(dst_wb_name)	
dst_wb.save(dst_wb_name)