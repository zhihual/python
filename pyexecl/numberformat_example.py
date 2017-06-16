import datetime
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
# set date using a Python datetime
ws['A1'] = datetime.datetime(2010, 7, 21)

ws['A1'].number_format
# You can enable type inference on a case-by-case basis
wb.guess_types = True
# set percentage using a string followed by the percent sign
ws['B1'] = '3.14%'
wb.guess_types = False
print ws['B1'].value
#0.031400000000000004

print ws['B1'].number_format
#'0%'