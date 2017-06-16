# -*- coding: UTF-8 -*-

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

wb = Workbook() # create one file

dest_filename = 'empty_book.xlsx' # set file name

ws1 = wb.active # 
ws1.title = "range names" # first sheet name

for row in range(1, 40): # 行
    ws1.append(range(600)) # 列

#range is real 1 ~ 40 eg

ws2 = wb.create_sheet(title="Pi")
ws2['F5'] = 3.14
ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col))) #取名字
print(ws3['AA10'].value)
#AA
wb.save(filename = dest_filename)