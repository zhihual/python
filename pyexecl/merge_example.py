from openpyxl.workbook import Workbook

wb = Workbook()
ws = wb.active

ws.merge_cells('A1:B1')
ws.unmerge_cells('A1:B1')
# or
#ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=4)
#ws.unmerge_cells(start_row=2,start_column=1,end_row=2,end_column=4)