#!/usr/bin/env python3
# -*- coding:utf-8 -*-

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl import worksheet
import os
import sys

reload(sys)
sys.setdefaultencoding('utf-8')

#  1. ��ʲô?  ���������λ����Ϣ������λ�ö�Ӧ������ 
#  2. ��ʲô�ô���  ������ʾ���л����µ��ź����Ʒֲ����ҵ��źŻ�۵㡣���ȵ㡣
#                   ����python�����̶�
#  3. ��ô����
#            a. ��ȡ������� �ֹ����� execl��� �����Ҫ�أ� ��źͻ�ַ. ������Ա�����ѯ
#            b. ����������ͼ�ϣ�ת��mac��ַΪ��Ÿ�ʽ��ͬ������� ��ѯexcel�� ͬ���������ַ
#
#
#
#


def openWhiteList(excelName):  #excelname �������excel���,��������handle
    print(excelName)
    wb = load_workbook(excelName)
    ws = wb.active
    row_num = len(tuple(ws.rows))
    print('row_num', row_num)
    for i in range (1, row_num):
        ws['C'+str(i)].value = 0
    return wb


def queryMeterAddr(excelhdl, meterId):#���ݵ��id,�ڵ��Ӱ������в�ѯ��ַ, ���ҽ�C��ֵ��1
    row_num = len(tuple(excelhdl.rows))
    columns_num = len(tuple(excelhdl.columns))
    #print(row_num, columns_num)
    for idx in range(1,row_num+1):
        #print(excelhdl['A'+str(idx)].value, excelhdl['B'+str(idx)].value)
        val = str(excelhdl['A'+str(idx)].value)
        val = val.strip()
        if val == meterId:
            val = str(excelhdl['B'+str(idx)].value)
            excelhdl['C'+str(idx)].value = 1 
            return val
    return ''

#def queryMeterId(excelhdl, meterAddr):
#    print('hello')
    

def convertMacToMeterId(mac):#��mac��ַת��Ϊ��id, return meterId
    mac = mac.strip()
    mac = mac.lstrip('\t')
    mac = mac.rstrip('\r\n')
    mac = mac.replace(':', '')
    #print(mac)
    val = mac[10:12]+mac[8:10]+mac[6:8]+mac[4:6]+mac[2:4]+mac[0:2]
    #print(val)
    return val
    

#def convertMeterIdToMac(meterId):#Convert meterId to Mac style. return Mac   


def queryOfflineMeter(excelhdl,offline_list):
    online_num = 0
    offline_num = 0
    row_num = len(tuple(excelhdl.rows))
    for i in range(1, row_num+1):
        local = excelhdl['C'+str(i)].value
        if local == 1:
            online_num +=1
        else:
            offline_num+=1
            temp_str = str(excelhdl['A'+str(i)].value)
            temp_str += '    '
            temp_str +=str(excelhdl['B'+str(i)].value)
            offline_list.append(temp_str)
    #print('Last', i)
    return (online_num, offline_num, row_num)

def openTopoTxt(topoTxtName):#topo txt file name, return handle
    print(topoTxtName)
    f = fopen(topoTxtName)
    return f


def putInfobyMac(topohdl, infoStr, placeMark):#output infoStr to topo table
    print('hello')





#Test MeterId

def Test_QueryMeterAddr():  
    tMeterId = '110120061848'
    tMeterAddr = queryMeterAddr(whiteListSheet, tMeterId)
    print(tMeterId, tMeterAddr.decode('gbk'))

#Test Convert Meter Mac

def Test_MacToMeterId():
    tMeterMac = '18:77:06:20:01:11'
    tConverMeterId = convertMacToMeterId(tMeterMac)
    print(tMeterMac)
    print(tConverMeterId)
#begin of main function

#main()
whitelListName='whitelist.xlsx'

whiteList = Workbook() 

topoTxtName='topo.txt'

whiteList= openWhiteList(whitelListName)

whiteListSheet = whiteList.active

print(len(tuple(whiteListSheet.rows)), len(tuple(whiteListSheet.columns)))

topo_list_content = []

offline_list = []

fh=open(topoTxtName, 'rb')

for line in fh.readlines():
    topo_list_content.append(line)
fh.close()

result_content = ''
for line in topo_list_content:
    if(line.find(':') >= 0):
        meter_id = convertMacToMeterId(line)
        addr = queryMeterAddr(whiteListSheet, meter_id)
        result_content += line[:-3] +'    '+ addr +'\r\n'

(online, offline, total) = queryOfflineMeter(whiteListSheet, offline_list)

print('Online:', online,'Offline:', offline, 'Total:', total)

offline_result_content = ''
offline_result_content += ('Online:'+str(online)+'  Offline:'+str(offline)+'  Total:'+str(total)+'\r\n')
result_content+=('Online:'+str(online)+'  Offline:'+str(offline)+'  Total:'+str(total)+'\r\n')

for line in offline_list:
    result_content+=line+'\r\n'
    offline_result_content+=line+'\r\n'
        
open('result.txt', 'wb').writelines(result_content)
open('offline_result.txt', 'wb').writelines(offline_result_content)













