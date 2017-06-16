# -*- coding: UTF-8 -*-



# Learn to use python to operate excel
from openpyxl import Workbook 

wb = Workbook() # delare one object

ws = wb.active  # grab the active worksheet

ws['A1'] = 42   # assign cell value directly

ws.append([1,2,3,4]) # Rows appened

import datetime 
ws['A2'] = datetime.datetime.now() # types will convert 

import pdb
pdb.set_trace()  #set running pdb breakpoint

wb.save("test.xlsx") # save file